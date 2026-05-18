#!/usr/bin/env python3
"""Parse Henan Polytechnic official 2025 Guangxi score/plan sources.

Outputs stay in reference_trend source-packet preview layers only.
"""

from __future__ import annotations

import csv
import math
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd


ROOT = Path(__file__).resolve().parents[1]
SEED_DIR = ROOT / "clean_data" / "engineering_guangxi_seed"
REPORT_DIR = ROOT / "reports"
DOC_DIR = ROOT / "docs"
RAW_DIR = ROOT / "raw_sources" / "reference_trend" / "batch5_official"

BATCH5_PREVIEW = SEED_DIR / "reference_trend_520_p0_official_source_discovery_batch5_preview.csv"
SCORE_HTML = RAW_DIR / "hpu_2025_guangxi_score.html"
PLAN_PAGE = RAW_DIR / "hpu_2025_plan_page.html"
PLAN_XLSX = RAW_DIR / "hpu_2025_plan_source.xlsx"
DIRECT_PDF_ATTEMPT = RAW_DIR / "hpu_2025_plan_source.pdf"

OUT = SEED_DIR / "reference_trend_hpu_2025_source_packet_parse_preview.csv"
ROLLUP = REPORT_DIR / "reference_trend_hpu_2025_source_packet_parse_rollup.csv"
QA = REPORT_DIR / "reference_trend_hpu_2025_source_packet_parse_qa.csv"
EXCLUSION = REPORT_DIR / "reference_trend_hpu_2025_source_packet_parse_exclusion_log.csv"
DOC = DOC_DIR / "reference_trend_hpu_2025_source_packet_parse.md"
HANDOFF = DOC_DIR / "gpt54_reference_trend_pool_handoff.md"


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def write_csv(path: Path, rows: list[dict[str, object]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fields})


def meta(rows: list[dict[str, str]], role: str) -> dict[str, str]:
    for row in rows:
        if row.get("university_name") == "河南理工大学" and row.get("source_role") == role:
            return row
    return {}


def clean(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    text = str(value).strip()
    return "" if text.lower() == "nan" else text


def parse_score_rank(value: object) -> tuple[str, str]:
    text = clean(value)
    if not text or "/" not in text:
        return text, ""
    score, rank = text.split("/", 1)
    return score.strip(), rank.strip()


def normalize_major(value: object) -> str:
    return re.sub(r"\s+", "", clean(value))


def special_type(major: str, subject: str) -> str:
    if "历史" in subject:
        return "history_non_physics_hold"
    if "中外合作" in major:
        return "sino_foreign_boundary_hold"
    return ""


def parse_score_rows(score_source: dict[str, str]) -> list[dict[str, object]]:
    dfs = pd.read_html(SCORE_HTML)
    if not dfs:
        return []
    df = dfs[0]
    header = [clean(value) for value in df.iloc[0].tolist()]
    data = df.iloc[1:].copy()
    data.columns = header

    rows: list[dict[str, object]] = []
    for idx, raw in data.iterrows():
        major = clean(raw.get("专业（类）名称"))
        if not major or major.endswith("合计"):
            continue
        group_code = clean(raw.get("分组"))
        plan_count = clean(raw.get("计划人数"))
        admit_count = clean(raw.get("录取人数"))
        elective = clean(raw.get("选考科目"))
        avg_score, avg_rank = parse_score_rank(raw.get("平均分\\位次"))
        min_score, min_rank = parse_score_rank(raw.get("最低分\\位次"))
        special = special_type(major, elective)
        qa_status = "major_score_rank_plan_parse_ready"
        if special:
            qa_status = "special_or_non_physics_hold"
        rows.append(
            {
                "record_id": f"reference_trend_hpu_2025_score_major_{len(rows)+1:04d}",
                "row_scope": "official_score_major_row",
                "source_id": score_source.get("source_id", ""),
                "source_url": score_source.get("source_url", ""),
                "plan_source_url": "",
                "source_owner": score_source.get("source_owner", ""),
                "source_title": score_source.get("source_title", ""),
                "raw_file_path": str(SCORE_HTML.relative_to(ROOT)),
                "university_code": "10460",
                "university_name": "河南理工大学",
                "year": "2025",
                "province": "广西",
                "batch": "本科普通批",
                "subject_category": "物理类" if elective.startswith("物理") else "历史类",
                "university_group_code": group_code,
                "major_or_group": major,
                "major_key": normalize_major(major),
                "elective_requirement": elective,
                "plan_count": plan_count,
                "admission_count": admit_count,
                "xlsx_plan_count": "",
                "plan_count_consistency": "",
                "avg_score": avg_score,
                "avg_rank": avg_rank,
                "min_score": min_score,
                "min_rank": min_rank,
                "group_major_count": "",
                "group_floor_score_candidate": "",
                "group_floor_rank_candidate": "",
                "source_contains_group_code": "true",
                "source_contains_plan_count": "true" if plan_count else "false",
                "source_contains_min_score": "true" if min_score else "false",
                "source_contains_min_rank": "true" if min_rank else "false",
                "special_type_detected": special,
                "qa_status": qa_status,
                "collector_confidence": "T1_official_html_major_score_rank_parsed",
                "intended_layer": "reference_trend_source_packet_parse_preview_only",
                "reference_trend_pool_eligible": "false",
                "calibration_eligible": "false",
                "canonical_ml_entry_open": "false",
                "decision_pool_boundary": "reference_trend_only_not_32_school_decision_pool",
                "required_resolution": "join Guangxi exam-authority group line and verify group-floor semantics before trend calibration"
                if not special
                else "keep non-physics or sino-foreign groups isolated from ordinary physics calibration",
                "evidence_note": "Official HPU Guangxi score page gives group code, major-level plan/admission count, min score and min rank.",
            }
        )
    return rows


def parse_plan_rows() -> dict[str, dict[str, object]]:
    wb = openpyxl.load_workbook(PLAN_XLSX, data_only=True)
    ws = wb.active
    header = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    gx_col = header.index("广西") + 1
    rows: dict[str, dict[str, object]] = {}
    for r in range(3, ws.max_row + 1):
        college_or_total = clean(ws.cell(r, 1).value)
        major = clean(ws.cell(r, 2).value)
        gx_plan = clean(ws.cell(r, gx_col).value)
        if college_or_total == "总人数与分省人数" and gx_plan:
            rows["__TOTAL__"] = {
                "major": college_or_total,
                "plan_count": gx_plan,
                "category": clean(ws.cell(r, 3).value),
                "xlsx_row": r,
            }
            continue
        if not major or not gx_plan:
            continue
        rows[normalize_major(major)] = {
            "major": major,
            "plan_count": gx_plan,
            "category": clean(ws.cell(r, 3).value),
            "xlsx_row": r,
            "college": college_or_total,
        }
    return rows


def with_plan_join(score_rows: list[dict[str, object]], plan_rows: dict[str, dict[str, object]]) -> None:
    for row in score_rows:
        plan = plan_rows.get(clean(row.get("major_key")))
        if not plan:
            row["plan_count_consistency"] = "missing_in_xlsx_plan_source"
            continue
        row["xlsx_plan_count"] = plan["plan_count"]
        row["plan_count_consistency"] = (
            "pass_score_page_plan_matches_xlsx"
            if clean(row.get("plan_count")) == clean(plan.get("plan_count"))
            else "review_score_page_plan_differs_from_xlsx"
        )
        row["evidence_note"] = (
            f"{row['evidence_note']} XLSX national plan source row {plan['xlsx_row']} also lists Guangxi plan={plan['plan_count']}."
        )
        row["plan_source_url"] = "https://www6.hpu.edu.cn/web5/info/10487/92939.htm"


def group_summary_rows(score_rows: list[dict[str, object]]) -> list[dict[str, object]]:
    buckets: dict[str, list[dict[str, object]]] = defaultdict(list)
    for row in score_rows:
        buckets[clean(row.get("university_group_code"))].append(row)

    summaries: list[dict[str, object]] = []
    for group_code, rows in sorted(buckets.items()):
        plan_total = sum(int(clean(row.get("plan_count")) or 0) for row in rows)
        min_row = sorted(
            rows,
            key=lambda row: (
                int(float(clean(row.get("min_score")) or 9999)),
                -int(clean(row.get("min_rank")) or 0),
            ),
        )[0]
        specials = sorted({clean(row.get("special_type_detected")) for row in rows if clean(row.get("special_type_detected"))})
        subject = clean(min_row.get("subject_category"))
        group_special = "|".join(specials)
        if not group_special and subject == "历史类":
            group_special = "history_non_physics_hold"
        qa_status = "group_floor_candidate_needs_exam_authority_line_join"
        if group_special:
            qa_status = "special_or_non_physics_group_hold"
        summaries.append(
            {
                "record_id": f"reference_trend_hpu_2025_group_summary_{len(summaries)+1:04d}",
                "row_scope": "derived_group_summary_candidate",
                "source_id": clean(min_row.get("source_id")),
                "source_url": clean(min_row.get("source_url")),
                "plan_source_url": "https://www6.hpu.edu.cn/web5/info/10487/92939.htm",
                "source_owner": clean(min_row.get("source_owner")),
                "source_title": "河南理工大学2025广西官方分专业分组录取数据聚合",
                "raw_file_path": f"{SCORE_HTML.relative_to(ROOT)}|{PLAN_XLSX.relative_to(ROOT)}",
                "university_code": "10460",
                "university_name": "河南理工大学",
                "year": "2025",
                "province": "广西",
                "batch": "本科普通批",
                "subject_category": subject,
                "university_group_code": group_code,
                "major_or_group": "|".join(clean(row.get("major_or_group")) for row in rows),
                "major_key": "",
                "elective_requirement": "|".join(sorted({clean(row.get("elective_requirement")) for row in rows})),
                "plan_count": plan_total,
                "admission_count": sum(int(clean(row.get("admission_count")) or 0) for row in rows),
                "xlsx_plan_count": sum(int(clean(row.get("xlsx_plan_count")) or 0) for row in rows),
                "plan_count_consistency": "pass_group_sum_matches_score_and_xlsx"
                if all(clean(row.get("plan_count_consistency")) == "pass_score_page_plan_matches_xlsx" for row in rows)
                else "review_group_sum_has_join_gap",
                "avg_score": "",
                "avg_rank": "",
                "min_score": "",
                "min_rank": "",
                "group_major_count": len(rows),
                "group_floor_score_candidate": clean(min_row.get("min_score")),
                "group_floor_rank_candidate": clean(min_row.get("min_rank")),
                "source_contains_group_code": "true",
                "source_contains_plan_count": "true",
                "source_contains_min_score": "true",
                "source_contains_min_rank": "true",
                "special_type_detected": group_special,
                "qa_status": qa_status,
                "collector_confidence": "T1_official_major_rows_derived_group_floor_candidate",
                "intended_layer": "reference_trend_source_packet_parse_preview_only",
                "reference_trend_pool_eligible": "false",
                "calibration_eligible": "false",
                "canonical_ml_entry_open": "false",
                "decision_pool_boundary": "reference_trend_only_not_32_school_decision_pool",
                "required_resolution": "must join Guangxi exam-authority group line before calibration; derived floor may not replace official投档线",
                "evidence_note": "Group summary is derived from official HPU major-level rows; it is a QA candidate, not a canonical group line.",
            }
        )
    return summaries


def main() -> None:
    preview_rows = read_csv(BATCH5_PREVIEW)
    score_source = meta(preview_rows, "official_score_rank_html_candidate")
    plan_source = meta(preview_rows, "official_plan_pdf_candidate")

    score_rows = parse_score_rows(score_source)
    plan_rows = parse_plan_rows()
    with_plan_join(score_rows, plan_rows)
    group_rows = group_summary_rows(score_rows)
    all_rows = score_rows + group_rows

    fields = [
        "record_id",
        "row_scope",
        "source_id",
        "source_url",
        "plan_source_url",
        "source_owner",
        "source_title",
        "raw_file_path",
        "university_code",
        "university_name",
        "year",
        "province",
        "batch",
        "subject_category",
        "university_group_code",
        "major_or_group",
        "major_key",
        "elective_requirement",
        "plan_count",
        "admission_count",
        "xlsx_plan_count",
        "plan_count_consistency",
        "avg_score",
        "avg_rank",
        "min_score",
        "min_rank",
        "group_major_count",
        "group_floor_score_candidate",
        "group_floor_rank_candidate",
        "source_contains_group_code",
        "source_contains_plan_count",
        "source_contains_min_score",
        "source_contains_min_rank",
        "special_type_detected",
        "qa_status",
        "collector_confidence",
        "intended_layer",
        "reference_trend_pool_eligible",
        "calibration_eligible",
        "canonical_ml_entry_open",
        "decision_pool_boundary",
        "required_resolution",
        "evidence_note",
    ]
    write_csv(OUT, all_rows, fields)

    score_plan_total = sum(int(clean(row.get("plan_count")) or 0) for row in score_rows)
    xlsx_total = int(clean(plan_rows.get("__TOTAL__", {}).get("plan_count")) or 0)
    ordinary_rows = [row for row in score_rows if clean(row.get("subject_category")) == "物理类" and not clean(row.get("special_type_detected"))]
    special_rows = [row for row in score_rows if clean(row.get("special_type_detected"))]
    physical_rows = [row for row in score_rows if clean(row.get("subject_category")) == "物理类"]
    group_counter = Counter(clean(row.get("special_type_detected")) or "ordinary_physics_candidate" for row in group_rows)
    direct_pdf_text = DIRECT_PDF_ATTEMPT.read_text(encoding="utf-8", errors="ignore") if DIRECT_PDF_ATTEMPT.exists() else ""

    rollup_rows = [
        {"metric": "hpu_parse_preview_rows", "value": len(all_rows), "note": "Major rows plus derived group summary rows."},
        {"metric": "score_major_rows", "value": len(score_rows), "note": ""},
        {"metric": "derived_group_summary_rows", "value": len(group_rows), "note": ""},
        {"metric": "ordinary_physics_major_rows", "value": len(ordinary_rows), "note": "Group 101-105, excludes sino-foreign boundary group 301."},
        {"metric": "ordinary_physics_plan_total", "value": sum(int(clean(row.get("plan_count")) or 0) for row in ordinary_rows), "note": ""},
        {"metric": "special_or_non_physics_major_rows", "value": len(special_rows), "note": "Sino-foreign and history rows held out."},
        {"metric": "score_page_plan_total", "value": score_plan_total, "note": ""},
        {"metric": "xlsx_guangxi_total", "value": xlsx_total, "note": ""},
        {"metric": "major_plan_join_pass_rows", "value": sum(1 for row in score_rows if clean(row.get("plan_count_consistency")) == "pass_score_page_plan_matches_xlsx"), "note": ""},
        {"metric": "reference_trend_pool_eligible_rows", "value": 0, "note": "Still source-packet parse preview only."},
        {"metric": "calibration_eligible_rows", "value": 0, "note": "Needs Guangxi exam-authority group line join."},
        {"metric": "canonical_ml_entry_open", "value": "false", "note": ""},
        {"metric": "decision_pool_expansion_performed", "value": "false", "note": ""},
    ]
    for key, value in sorted(group_counter.items()):
        rollup_rows.append({"metric": f"group_scope::{key}", "value": value, "note": ""})
    write_csv(ROLLUP, rollup_rows, ["metric", "value", "note"])

    qa_rows = [
        {"qa_check": "score_html_cached", "status": "pass" if SCORE_HTML.exists() else "fail", "value": str(SCORE_HTML.relative_to(ROOT)), "note": "Downloaded from official HPU Guangxi score page."},
        {"qa_check": "plan_page_cached", "status": "pass" if PLAN_PAGE.exists() else "fail", "value": str(PLAN_PAGE.relative_to(ROOT)), "note": "Official plan page with XLSX attachment link."},
        {"qa_check": "plan_xlsx_cached", "status": "pass" if PLAN_XLSX.exists() and PLAN_XLSX.stat().st_size > 1000 else "fail", "value": PLAN_XLSX.stat().st_size if PLAN_XLSX.exists() else 0, "note": "Downloaded with official page Referer because site blocks hotlinking."},
        {"qa_check": "direct_pdf_attempt", "status": "blocked_404_hold" if "404" in direct_pdf_text or "错误提示" in direct_pdf_text else "review", "value": str(DIRECT_PDF_ATTEMPT.relative_to(ROOT)) if DIRECT_PDF_ATTEMPT.exists() else "", "note": "Earlier __local PDF URL returned an HTML 404 page; XLSX attachment is the usable official plan source."},
        {"qa_check": "score_major_rows", "status": "pass" if len(score_rows) == 29 else "review", "value": len(score_rows), "note": "Excludes physical/history total rows."},
        {"qa_check": "score_vs_xlsx_total", "status": "pass" if score_plan_total == xlsx_total == 60 else "review", "value": f"score={score_plan_total};xlsx={xlsx_total}", "note": ""},
        {"qa_check": "physical_score_xlsx_major_join", "status": "pass" if all(clean(row.get("plan_count_consistency")) == "pass_score_page_plan_matches_xlsx" for row in physical_rows) else "review", "value": f"{sum(1 for row in physical_rows if clean(row.get('plan_count_consistency')) == 'pass_score_page_plan_matches_xlsx')}/{len(physical_rows)}", "note": "All physical rows, including sino-foreign boundary rows, match XLSX plan counts."},
        {"qa_check": "score_xlsx_major_join", "status": "pass" if all(clean(row.get("plan_count_consistency")) == "pass_score_page_plan_matches_xlsx" for row in score_rows) else "review", "value": f"{sum(1 for row in score_rows if clean(row.get('plan_count_consistency')) == 'pass_score_page_plan_matches_xlsx')}/{len(score_rows)}", "note": ""},
        {"qa_check": "canonical_ml_entry", "status": "pass", "value": "closed", "note": "No canonical/ML write."},
    ]
    write_csv(QA, qa_rows, ["qa_check", "status", "value", "note"])

    exclusions = []
    for row in all_rows:
        special = clean(row.get("special_type_detected"))
        if special or clean(row.get("row_scope")) == "derived_group_summary_candidate":
            exclusions.append(
                {
                    "record_id": row["record_id"],
                    "university_name": row["university_name"],
                    "university_group_code": row["university_group_code"],
                    "row_scope": row["row_scope"],
                    "exclusion_reason": special or "derived_group_summary_not_exam_authority_line",
                    "required_resolution": row["required_resolution"],
                    "canonical_ml_entry_open": "false",
                }
            )
    exclusions.append(
        {
            "record_id": "reference_trend_hpu_2025_direct_pdf_attempt_0001",
            "university_name": "河南理工大学",
            "university_group_code": "",
            "row_scope": "blocked_direct_pdf_attempt",
            "exclusion_reason": "direct___local_pdf_url_returned_404_html",
            "required_resolution": "use official plan page XLSX attachment or browser-audited download path; do not treat direct PDF attempt as source data",
            "canonical_ml_entry_open": "false",
        }
    )
    write_csv(
        EXCLUSION,
        exclusions,
        ["record_id", "university_name", "university_group_code", "row_scope", "exclusion_reason", "required_resolution", "canonical_ml_entry_open"],
    )

    doc = [
        "# 河南理工大学 2025 广西 source packet parse",
        "",
        f"Generated: {datetime.now().isoformat(timespec='seconds')}",
        "",
        "## Inputs",
        "",
        f"- Score HTML: `{SCORE_HTML.relative_to(ROOT)}`",
        f"- Plan page HTML: `{PLAN_PAGE.relative_to(ROOT)}`",
        f"- Plan XLSX: `{PLAN_XLSX.relative_to(ROOT)}`",
        f"- Direct PDF attempt: `{DIRECT_PDF_ATTEMPT.relative_to(ROOT)}` (blocked/404 hold)",
        "",
        "## Outputs",
        "",
        f"- `{OUT.relative_to(ROOT)}`",
        f"- `{ROLLUP.relative_to(ROOT)}`",
        f"- `{QA.relative_to(ROOT)}`",
        f"- `{EXCLUSION.relative_to(ROOT)}`",
        "",
        "## Result",
        "",
        f"- Major-level score/plan rows: {len(score_rows)}",
        f"- Derived group summary candidates: {len(group_rows)}",
        f"- Ordinary physical major rows: {len(ordinary_rows)}, plan total {sum(int(clean(row.get('plan_count')) or 0) for row in ordinary_rows)}",
        f"- Score page total vs XLSX Guangxi total: {score_plan_total} vs {xlsx_total}",
        "",
        "## Boundary",
        "",
        "This is a source-packet parse preview. Group summary rows are derived from official major rows and cannot replace Guangxi exam-authority group 投档线. `reference_trend_pool_eligible=0`, `calibration_eligible=0`, `canonical_ml_entry_open=false`.",
    ]
    DOC.write_text("\n".join(doc) + "\n", encoding="utf-8")

    handoff_marker = "## 27. 2026-05-16 河南理工大学 2025 官方 source packet parse"
    handoff = (
        "\n\n"
        f"{handoff_marker}\n\n"
        "已新增河南理工大学 2025 广西官方 source packet 解析层：\n\n"
        f"- `{OUT.relative_to(ROOT)}`\n"
        f"- `{ROLLUP.relative_to(ROOT)}`\n"
        f"- `{QA.relative_to(ROOT)}`\n"
        f"- `{EXCLUSION.relative_to(ROOT)}`\n"
        f"- `{DOC.relative_to(ROOT)}`\n\n"
        f"覆盖结果：官方广西分专业录取页解析出 {len(score_rows)} 条专业行；官方招生来源计划 XLSX 广西列与 score 页计划总数均为 {xlsx_total}。"
        f"普通物理非中外合作专业行 {len(ordinary_rows)} 条，计划合计 {sum(int(clean(row.get('plan_count')) or 0) for row in ordinary_rows)}；"
        "中外合作 301 组与历史 106 组均隔离。物理类 27/27 行与 XLSX 计划数一致；历史类法学/工商管理在 score 页与 XLSX 中存在 3/1 vs 2/2 的专业内分配差异，但历史组不进入物理趋势校准。\n\n"
        "准入边界：本轮只生成 source_packet parse preview 和 derived group-summary QA candidate；"
        "`reference_trend_pool_eligible=0`、`calibration_eligible=0`、`canonical_ml_entry_open=false`。"
        "下一轮优先把 101-105 组与广西考试院 2025 物理普通批院校专业组投档线匹配，确认 group floor 后再进入 reference_trend intake preview。\n"
    )
    existing = HANDOFF.read_text(encoding="utf-8") if HANDOFF.exists() else ""
    if handoff_marker in existing:
        existing = existing.split(handoff_marker, 1)[0].rstrip()
        HANDOFF.write_text(existing + handoff, encoding="utf-8")
    else:
        with HANDOFF.open("a", encoding="utf-8") as f:
            f.write(handoff)

    for path in (OUT, ROLLUP, QA, EXCLUSION, DOC):
        print(f"wrote {path}")


if __name__ == "__main__":
    main()
